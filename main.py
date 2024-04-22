from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook

import time

def parse(driver):
    scroll(driver, 30, 1)
    product_cards = driver.find_elements(By.CLASS_NAME, 'product-card')
    items = []

    for card in product_cards:
        link = card.find_element(By.CSS_SELECTOR, 'a.product-card__link').get_attribute('href')
        title = card.find_element(By.CLASS_NAME, 'product-card__name').text
        price = card.find_element(By.CLASS_NAME, 'price__lower-price').text
        brand = card.find_element(By.CLASS_NAME, 'product-card__brand').text
        rating = card.find_element(By.CLASS_NAME, 'address-rate-mini').text
        img = card.find_element(By.CLASS_NAME, 'j-thumbnail').get_attribute('src')
        items.append([link, title, price, brand, rating, img])

    return items

def scroll(driver, times, delay):
    for i in range(times):  
        driver.execute_script("window.scrollBy(0, 400);")
        time.sleep(delay)
    

def next_page(driver):
    next_page_element = driver.find_element(By.CLASS_NAME, 'pagination-next')
    url = next_page_element.get_attribute('href')
    driver.get(url)

def write_to_excel(items):
    try:
        wb = load_workbook("creatine.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    for row in items:
        ws.append(row)

    wb.save("creatine.xlsx")

def main():
    url = 'https://www.wildberries.ru/catalog/0/search.aspx?search=%D0%BA%D1%80%D0%B5%D0%B0%D1%82%D0%B8%D0%BD#c5544385'
    driver_path = "C:\\Program Files\\ChromeDriver\\chromedriver.exe"
    chrome_binary_path = "C:\\Users\\user\\Downloads\\GoogleChromePortableBeta\\App\\Chrome-bin\\chrome.exe"
    
    options = Options()
    options.add_experimental_option("detach", True)
    options.binary_location = chrome_binary_path
    
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    driver.get(url)
    
    pages = 5
    header_items = [['Ссылка', 'Название', 'Цена', 'Производитель', 'Рейтинг', 'Картинка']]
    write_to_excel(header_items)
    for _ in range(pages):
        items = parse(driver)
        write_to_excel(items)
        next_page(driver)
        
if __name__ == "__main__":
    main()
